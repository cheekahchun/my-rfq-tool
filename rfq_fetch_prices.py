"""
RFQ Price Fetcher
Logs into the internal RFQ site, finds the specified RFQ, checks if it has a '_Done.xlsx' file,
downloads it, and extracts the cost prices.
"""

import os
import re
import requests
import openpyxl
from io import BytesIO

# Config reuse
base = os.path.dirname(os.path.abspath(__file__))
config_path = os.path.join(base, "config.txt")
_config = {}
if os.path.exists(config_path):
    with open(config_path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if "=" in line and not line.startswith("#"):
                k, v = line.split("=", 1)
                _config[k.strip()] = v.strip()

RFQ_BASE_URL = _config.get("RFQ_URL", "http://192.168.68.33")
RFQ_EMAIL    = _config.get("RFQ_EMAIL", "kcchee@genxai.com.my")
RFQ_PASSWORD = _config.get("RFQ_PASSWORD", "12345678")
LOGIN_URL    = f"{RFQ_BASE_URL}/genxai/auth/login"

def _login(session: requests.Session) -> str | None:
    try:
        session.headers.update({"User-Agent": "Mozilla/5.0"})
        resp = session.get(LOGIN_URL, timeout=10)
        csrf_match = re.search(r'<meta name=["\']csrf-token["\'] content=["\']([^"\']+)["\']', resp.text)
        csrf_token = csrf_match.group(1) if csrf_match else ""

        login_resp = session.post(
            LOGIN_URL,
            data={"email": RFQ_EMAIL, "password": RFQ_PASSWORD, "_token": csrf_token},
            allow_redirects=True, timeout=10
        )
        if "auth/login" in login_resp.url:
            return "RFQ login failed. Check config credentials."
        return None
    except Exception as e:
        return f"Cannot reach RFQ site: {e}"

def fetch_rfq_cost(rfq_no: str) -> dict:
    """
    Attempts to download the _Done.xlsx file for a given RFQ
    and calculate the total cost.
    
    Returns:
    {
        "success": bool,
        "msg": str,
        "total_cost": float,
        "currency": "MYR",
        "file_found": bool
    }
    """
    session = requests.Session()
    err = _login(session)
    if err:
        return {"success": False, "msg": err, "total_cost": 0, "file_found": False}

    # 1. Search for the RFQ from the first few pages of the list
    try:
        found_detail_url = None
        for p_idx in range(1, 6): # Try first 5 pages
            list_url = f"{RFQ_BASE_URL}/genxai/rfqs"
            if p_idx > 1: list_url += f"?page={p_idx}"
            
            list_resp = session.get(list_url, timeout=10)
            html = list_resp.text
            
            # Find the row containing the RFQ No
            row_html = ""
            for row_str in html.split('<tr'):
                if rfq_no in row_str:
                    row_html = '<tr' + row_str
                    break
            
            if row_html:
                # Found the row, extract the detail link
                detail_match = re.search(fr'href="([^"]+/genxai/rfqs/\d+)"', row_html)
                if detail_match:
                    found_detail_url = detail_match.group(1)
                    break # Success!
                    
        if not found_detail_url:
            return {"success": False, "msg": f"Could not find RFQ {rfq_no} in the system (scanned 5 pages).", "total_cost": 0, "file_found": False}
            
        detail_url = found_detail_url
        
    except Exception as e:
         return {"success": False, "msg": f"Search error: {e}", "total_cost": 0, "file_found": False}
        
        
    except Exception as e:
         return {"success": False, "msg": f"Search error: {e}", "total_cost": 0, "file_found": False}


    # 2. Go to the Details page and look for the Done.xlsx file
    try:
        detail_resp = session.get(detail_url, timeout=10)
        html = detail_resp.text
        
        # Look for the download link matching _Done.xlsx or Done.xlsx
        download_match = re.search(r'href="([^"]*[Dd]one[^"]*\.xlsx)"', html)
        if not download_match:
            return {"success": True, "msg": f"RFQ {rfq_no} is found, but procurement has not attached the Done file yet.", "total_cost": 0, "file_found": False}
            
        download_url = download_match.group(1)
        
        # --- Extract Item Images ---
        images = []
        # Find quotation table section
        quot_idx = html.find("Quotation:")
        search_html = html[quot_idx:] if quot_idx != -1 else html
        
        img_matches = re.findall(r'<img[^>]+src="([^"]+)"', search_html)
        for img_path in img_matches:
            if 'logo' in img_path.lower() or 'avatar' in img_path.lower() or 'profile' in img_path.lower(): continue
            full_img_url = img_path if img_path.startswith('http') else RFQ_BASE_URL + img_path
            images.append(full_img_url)
            

    except Exception as e:
        return {"success": False, "msg": f"Edit page error: {e}", "total_cost": 0, "file_found": False}


    # 3. Download the Excel file
    try:
        file_resp = session.get(download_url, timeout=15)
        if file_resp.status_code != 200:
            return {"success": False, "msg": "Failed to download the Excel file.", "total_cost": 0, "file_found": True}
        
        excel_bytes = file_resp.content
        
        # 4. Parse the Excel file from memory
        wb = openpyxl.load_workbook(BytesIO(excel_bytes), data_only=True)
        ws = wb.active
        
        # Find column indices for QTY, Description and Cost
        headers = [str(c.value).strip().lower() for c in ws[1] if c.value]
        qty_idx = -1
        cost_idx = -1
        desc_idx = -1
        
        for i, h in enumerate(headers):
            if h == "qty":
                qty_idx = i
            elif "description" in h or "item" in h or "item description" in h:
                desc_idx = i
            elif "landed cost / unit" in h or "cost" in h:
                # E.g. "landed cost / unit (myr)"
                cost_idx = i
                
        if qty_idx == -1 or cost_idx == -1:
             return {"success": False, "msg": "Could not find 'QTY' or 'Cost' columns in the downloaded Excel.", "total_cost": 0, "file_found": True}
             
        # Calculate Total Cost = Sum(QTY * Unit Cost)
        total_cost = 0.0
        breakdown = []
        for row in ws.iter_rows(min_row=2, max_col=len(headers), values_only=True):
            if not any(row): continue # skip empty rows
            
            try:
                qty_val = row[qty_idx]
                cost_val = row[cost_idx]
                desc_val = row[desc_idx] if desc_idx != -1 else "Item"
                
                # Smart number extraction to handle things like "RM 120.00" or "15 pcs"
                def extract_float(val, default=0.0):
                    if val is None or str(val).strip() == "": return default
                    s = str(val).replace(',', '')
                    m = re.search(r'[-+]?\d*\.\d+|\d+', s)
                    return float(m.group()) if m else default

                qty = extract_float(qty_val, default=1.0)
                cost = extract_float(cost_val, default=0.0)
                subtotal = qty * cost
                
                if qty > 0 or cost > 0:
                    total_cost += subtotal
                    breakdown.append({
                        "desc": str(desc_val),
                        "qty": qty,
                        "cost": cost,
                        "subtotal": subtotal
                    })
            except Exception as e:
                pass # Ignore headers or completely blank rows
                
        wb.close()
        
        return {
            "success": True, 
            "msg": f"Successfully pulled price from Done file!", 
            "total_cost": round(total_cost, 2), 
            "breakdown": breakdown,
            "images": images,
            "file_found": True
        }
        
    except Exception as e:
        return {"success": False, "msg": f"Error parsing Excel file: {e}", "total_cost": 0, "file_found": True}

def inject_images_to_master(rfq_no, images, master_path):
    """Inject extracted portal images into Master_Orders.xlsx Customer's Ref column."""
    if not images: return False, "No images found."
    try:
        from openpyxl.drawing.image import Image as OpenpyxlImage
        from io import BytesIO
        import requests
        import openpyxl
        import openpyxl.utils
        
        wb = openpyxl.load_workbook(master_path)
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        
        uid_col = -1
        ref_col = -1
        for i, h in enumerate(headers):
            if h == 'uid' or h == 'ref rfq no.': uid_col = i + 1
            if 'reference' in h or "customer's ref" in h: ref_col = i + 1
            
        if uid_col == -1 or ref_col == -1: return False, "UID or Customer's Reference columns missing in Master_Orders"
        
        # Find all rows for this rfq
        target_rows = []
        for row in range(2, ws.max_row + 1):
            cell_val = str(ws.cell(row=row, column=uid_col).value)
            if cell_val == rfq_no:
                target_rows.append(row)
                
        # Inject images in order
        injected_count = 0
        for idx, row_num in enumerate(target_rows):
            if idx < len(images):
                img_url = images[idx]
                try:
                    resp = requests.get(img_url, timeout=10)
                    if resp.status_code == 200:
                        img_io = BytesIO(resp.content)
                        img = OpenpyxlImage(img_io)
                        # Resize image to fit nicely within the cell
                        img.width = 120
                        img.height = 120
                        # Auto-resize row and column to fit the image
                        ws.row_dimensions[row_num].height = 95
                        ws.column_dimensions[openpyxl.utils.get_column_letter(ref_col)].width = 22
                        
                        # Set cell value and insert image
                        cell_anchor = f"{openpyxl.utils.get_column_letter(ref_col)}{row_num}"
                        # Clear text just in case
                        ws.cell(row=row_num, column=ref_col).value = ""
                        ws.add_image(img, cell_anchor)
                        injected_count += 1
                except Exception as e:
                    print(f"Skipping Image {idx}: {e}")
                    
        wb.save(master_path)
        return True, f"Injected {injected_count} images"
    except Exception as e:
        return False, f"Image Injection Error: {str(e)}"

# For simple testing
if __name__ == "__main__":
    res = fetch_rfq_cost("GX001377")
    print(res)
