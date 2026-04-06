"""
RFQ Auto-Submitter
Logs into the internal RFQ site and submits an order's items as a new RFQ.
Credentials are read from config.txt (same as rfq_scraper.py).
"""

import os
import re
import io
import requests
import openpyxl
import datetime

# ------------------------------------------------------------------
# Config  (same pattern as rfq_scraper.py)
# ------------------------------------------------------------------
def _get_config():
    base = os.path.dirname(os.path.abspath(__file__))
    for candidate in [base, os.path.dirname(base)]:
        p = os.path.join(candidate, "config.txt")
        if os.path.exists(p):
            cfg = {}
            with open(p, encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if "=" in line and not line.startswith("#"):
                        k, v = line.split("=", 1)
                        cfg[k.strip()] = v.strip()
            return cfg
    return {}

_config      = _get_config()
RFQ_BASE_URL = _config.get("RFQ_URL",      "http://192.168.68.33")
RFQ_EMAIL    = _config.get("RFQ_EMAIL", "")     # <--- 从 config.txt 读取
RFQ_PASSWORD = _config.get("RFQ_PASSWORD", "")  # <--- 从 config.txt 读取

LOGIN_URL      = f"{RFQ_BASE_URL}/genxai/auth/login"
RFQ_CREATE_URL = f"{RFQ_BASE_URL}/genxai/rfqs"


# ------------------------------------------------------------------
# Excel template builder  (matches the green template the user uses)
# ------------------------------------------------------------------
TEMPLATE_HEADERS = [
    "Customer Name",
    "Item Name",
    "Description (P/N)",
    "QTY",
    "Customer's Reference (Choose from the Amazon Link)",
    "Link",
    "Accept Alternative",
    "Image",
    "Shipment Charges",
    "Landed Cost / Unit (MYR)",
    "Lead Time",
    "Extra Info (Purchaser)"
]

def _build_template_excel(order_row: dict) -> bytes:
    """
    Build an in-memory Excel file by loading the existing Template.xlsx
    to preserve all colors, formatting and headers.
    """
    base = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(base, "Template.xlsx")
    
    if os.path.exists(template_path):
        wb = openpyxl.load_workbook(template_path)
    else:
        wb = openpyxl.Workbook()
        
    ws = wb.active

    # 1. Capture styles and column widths from row 2 BEFORE doing anything
    from copy import copy
    row2_styles = []
    col_widths = {}
    if ws.max_row >= 2:
        for c in range(1, 14): # Covers the 12 columns + wiggle room
            cell = ws.cell(row=2, column=c)
            style_dict = {}
            if cell.has_style:
                style_dict['font'] = copy(cell.font)
                style_dict['border'] = copy(cell.border)
                style_dict['fill'] = copy(cell.fill)
                style_dict['number_format'] = copy(cell.number_format)
                style_dict['protection'] = copy(cell.protection)
                style_dict['alignment'] = copy(cell.alignment)
            row2_styles.append(style_dict)
            
            # Capture column width
            col_letter = cell.column_letter
            if col_letter in ws.column_dimensions:
                col_widths[col_letter] = copy(ws.column_dimensions[col_letter])

    # 2. Clear ALL values from row 2 downward (keeping the rows)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    def parse_bullets(text):
        if not text or str(text).lower() == 'nan': return []
        text = str(text)
        if '•' in text:
            return [x.strip() for x in text.split('•') if x.strip()]
        return [text.strip()]

    # Mapping Dashboard keys to the template data
    customer = str(order_row.get("Customer Name", "") or "")
    if customer.lower() == 'nan': customer = ""
    
    item_names = parse_bullets(order_row.get("Item Name", ""))
    descs = parse_bullets(order_row.get("Description", "")) # Use 'Description' from dashboard
    qtys = parse_bullets(order_row.get("QTY", ""))
    refs = parse_bullets(order_row.get("Customer's Ref", "")) # Use 'Customer's Ref' from dashboard
    links = parse_bullets(order_row.get("Link", ""))
    alts = parse_bullets(order_row.get("Accept Alternative", ""))

    num_items = max(len(item_names), 1)

    for i in range(num_items):
        r = i + 2
        
        # Only print Customer Name on the first row
        if i == 0:
            ws.cell(row=r, column=1, value=customer)
            
        ws.cell(row=r, column=2, value=item_names[i] if i < len(item_names) else "")
        ws.cell(row=r, column=3, value=descs[i] if i < len(descs) else "")
        
        qty_val = qtys[i] if i < len(qtys) else ""
        if str(qty_val).isdigit(): qty_val = int(qty_val)
        ws.cell(row=r, column=4, value=qty_val)
        
        ws.cell(row=r, column=5, value=refs[i] if i < len(refs) else "")
        ws.cell(row=r, column=6, value=links[i] if i < len(links) else "")
        ws.cell(row=r, column=7, value=alts[i] if i < len(alts) else "")
        
        # 3. Re-apply styles from the saved row2_styles
        for c_idx, s in enumerate(row2_styles, start=1):
            if not s: continue
            target_cell = ws.cell(row=r, column=c_idx)
            target_cell.font = s['font']
            target_cell.border = s['border']
            target_cell.fill = s['fill']
            target_cell.number_format = s['number_format']
            target_cell.protection = s['protection']
            target_cell.alignment = s['alignment']

    # 4. Re-apply column widths
    for letter, dim in col_widths.items():
        ws.column_dimensions[letter] = dim


    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generate_draft_excel(order_row: dict) -> str:
    """Generates the Excel draft and saves it to Extracted_Emails."""
    excel_bytes = _build_template_excel(order_row)
    base = os.path.dirname(os.path.abspath(__file__))
    extracted_dir = os.path.join(base, "Extracted_Emails")
    os.makedirs(extracted_dir, exist_ok=True)
    
    # Safe name based on Customer
    import re
    cust = re.sub(r'[\\/*?:"<>|]', "", str(order_row.get("Customer Name", "Customer")))
    filename = f"Draft_{cust}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(extracted_dir, filename)
    
    with open(filepath, 'wb') as f:
        f.write(excel_bytes)
    return filepath


# ------------------------------------------------------------------
# Login helper
# ------------------------------------------------------------------
def _login() -> tuple[requests.Session | None, str | None]:
    """Returns (logged-in session, error) tuple."""
    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0"})

    try:
        resp = session.get(LOGIN_URL, timeout=10)
    except Exception as e:
        return None, f"Cannot reach RFQ site: {e}"

    csrf_match = re.search(
        r'<meta name=["\']csrf-token["\'] content=["\']([^"\']+)["\']',
        resp.text,
    )
    csrf_token = csrf_match.group(1) if csrf_match else ""

    login_resp = session.post(
        LOGIN_URL,
        data={"email": RFQ_EMAIL, "password": RFQ_PASSWORD, "_token": csrf_token},
        allow_redirects=True,
        timeout=10,
    )

    if "auth/login" in login_resp.url:
        return None, "RFQ login failed — check RFQ_EMAIL / RFQ_PASSWORD in config.txt"

    return session, None


# ------------------------------------------------------------------
# Main public function
# ------------------------------------------------------------------
def submit_rfq(order_row: dict, rfq_no: str = "", item_desc: str = "", filepath: str = "") -> tuple[bool, str]:
    """
    Submit a new RFQ to the internal company portal.

    Args:
        order_row : dict with at least Customer Name, Item Name, Description, QTY
        rfq_no    : optional RFQ number override; auto-generated if empty
        item_desc : optional Item Description textarea text
        filepath  : optional local file path of a pre-verified draft Excel file to upload

    Returns:
        (success: bool, message: str)
    """
    # Auto-generate RFQ No if not provided
    if not rfq_no:
        rfq_no = f"GX-AUTO-{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}"

    if not item_desc:
        item_desc = f"{order_row.get('Item Name', '')} x{order_row.get('QTY', '')} — {order_row.get('Customer Name', '')}"

    # Login
    session, err = _login()
    if err:
        return False, err

    # Get CSRF token from create page
    try:
        create_resp = session.get(f"{RFQ_BASE_URL}/genxai/rfqs/create", timeout=10)
    except Exception as e:
        return False, f"Could not load RFQ create page: {e}"

    csrf_match = re.search(
        r'<meta name=["\']csrf-token["\'] content=["\']([^"\']+)["\']',
        create_resp.text,
    )
    csrf_token = csrf_match.group(1) if csrf_match else ""

    # Also look for _token inside the form
    form_token_match = re.search(r'name="_token"\s+value="([^"]+)"', create_resp.text)
    form_token = form_token_match.group(1) if form_token_match else csrf_token

    # Build Excel
    if filepath and os.path.exists(filepath):
        with open(filepath, 'rb') as f:
            excel_bytes = f.read()
        excel_filename = os.path.basename(filepath)
    else:
        excel_bytes = _build_template_excel(order_row)
        excel_filename = f"RFQ_{rfq_no.replace(' ', '_')}.xlsx"

    # Submit
    try:
        resp = session.post(
            RFQ_CREATE_URL,
            data={
                "_token":    form_token,
                "rfq_no":    rfq_no,
                "items_desc": item_desc,
            },
            files={
                "item_reference_file_path": (
                    excel_filename,
                    excel_bytes,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            allow_redirects=True,
            timeout=15,
        )
    except Exception as e:
        return False, f"Submission error: {e}"

    if resp.status_code in (200, 302) and "auth/login" not in resp.url:
        # Try to parse the actual RFQ No. assigned by the server from the response HTML
        assigned_rfq_match = re.search(r'GX\d+', resp.text)
        actual_rfq_no = assigned_rfq_match.group(0) if assigned_rfq_match else rfq_no

        # Log to Google Sheet (IF ENABLED)
        try:
            if _config.get("ENABLE_GSHEET", "True").upper() == "TRUE":
                import gsheet_logger
                gs_ok, gs_msg = gsheet_logger.log_rfq_submission(order_row, actual_rfq_no)
                if gs_ok:
                    return True, f"RFQ '{actual_rfq_no}' submitted + logged to Google Sheet!"
                else:
                    return True, f"RFQ '{actual_rfq_no}' submitted! (Google Sheet: {gs_msg})"
            else:
                return True, f"RFQ '{actual_rfq_no}' submitted successfully!"
        except Exception as gs_err:
            return True, f"RFQ '{actual_rfq_no}' submitted! (Google Sheet skipped: {gs_err})"
    else:
        return False, f"Submission may have failed (status {resp.status_code}). Please check the RFQ site."
