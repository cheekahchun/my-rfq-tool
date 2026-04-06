import streamlit as st
import pandas as pd
import os
import sys
import time
import re
from datetime import datetime
import auth_manager

# --- CONFIG & PATHS ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
MASTER_FILE = os.path.join(BASE_DIR, "Master_Orders.xlsx")
CONFIG_PATH = os.path.join(BASE_DIR, "config.txt")

# Set Page Config
st.set_page_config(page_title="AI Procurement Admin", page_icon="🛡️", layout="wide")

# Custom UI Styling
st.markdown("""
<style>
    .stApp { background: radial-gradient(circle at top right, #1a1c24, #0e1117); }
    [data-testid="stMetricValue"] { font-size: 2.2rem; color: #00d4ff; }
    .animated-title {
        background: linear-gradient(90deg, #00d4ff, #9d00ff, #00d4ff);
        background-size: 200% auto;
        color: #fff; background-clip: text; -webkit-background-clip: text; -webkit-text-fill-color: transparent;
        animation: gradient-text 3s linear infinite; font-size: 2.5rem; font-weight: 700;
    }
    @keyframes gradient-text { 0% { background-position: 0% 50%; } 50% { background-position: 100% 50%; } 100% { background-position: 0% 50%; } }
    .comm-icon {
        display: flex; align-items: center; justify-content: space-between;
        padding: 10px; border-radius: 10px; background: rgba(255,255,255,0.05); margin-bottom: 8px;
        transition: all 0.3s ease; border: 1px solid transparent;
        cursor: pointer; text-decoration: none; color: inherit;
    }
    .comm-icon:hover { background: rgba(255,255,255,0.1); border-color: rgba(0,212,255,0.3); transform: translateX(5px); color: #fff; }
    .comm-status { font-size: 0.7rem; padding: 2px 6px; border-radius: 5px; color: #fff; font-weight: bold; }
    .status-active { background: #00eb93; box-shadow: 0 0 10px rgba(0,235,147,0.4); }
    .status-check { background: #00d4ff; }
    .notif-dot {
        width: 12px; height: 12px; background-color: #00eb93; border-radius: 50%;
        display: inline-block; vertical-align: middle;
        box-shadow: 0 0 10px #00eb93; animation: pulse 2.5s infinite;
        margin-left: 10px;
    }
    @keyframes pulse {
        0% { transform: scale(0.85); box-shadow: 0 0 0 0 rgba(0, 235, 147, 0.8); }
        50% { transform: scale(1.2); box-shadow: 0 0 15px 5px rgba(0, 235, 147, 0.3); }
        100% { transform: scale(0.85); box-shadow: 0 0 0 0 rgba(0, 235, 147, 0); }
    }
    .action-card {
        background: rgba(255, 255, 255, 0.03); padding: 20px; border-radius: 12px;
        border: 1px solid rgba(255,255,255,0.05); min-height: 220px;
    }
    .info-box {
        background: rgba(0, 212, 255, 0.05); border-left: 4px solid #00d4ff;
        padding: 15px; border-radius: 0 10px 10px 0; margin-bottom: 20px;
    }
    /* Timeline Stying */
    .timeline-container { display: flex; justify-content: space-between; align-items: center; margin: 30px 10px; position: relative; }
    .timeline-line { position: absolute; top: 15px; left: 0; right: 0; height: 3px; background: rgba(255,255,255,0.1); z-index: 1; }
    .timeline-progress { position: absolute; top: 15px; left: 0; height: 3px; background: #00d4ff; z-index: 2; transition: width 0.8s ease-in-out; }
    .timeline-step { position: relative; z-index: 3; display: flex; flex-direction: column; align-items: center; width: 80px; }
    .timeline-dot { 
        width: 32px; height: 32px; border-radius: 50%; background: #1a1c24; border: 2px solid rgba(255,255,255,0.2);
        display: flex; align-items: center; justify-content: space-between; color: rgba(255,255,255,0.3); font-size: 14px;
        transition: all 0.5s ease;
        margin-bottom: 8px; justify-content: center;
    }
    .timeline-step.active .timeline-dot { border-color: #00d4ff; color: #00d4ff; box-shadow: 0 0 15px rgba(0,212,255,0.5); background: rgba(0,212,255,0.1); }
    .timeline-step.done .timeline-dot { background: #00eb93; border-color: #00eb93; color: #000; box-shadow: 0 0 10px rgba(0,235,147,0.3); }
    .timeline-label { font-size: 0.75rem; color: rgba(255,255,255,0.5); font-weight: 500; text-align: center; }
    .timeline-step.active .timeline-label { color: #fff; font-weight: 700; }
    .timeline-active-pulse { animation: timeline-pulse 2s infinite; }
    @keyframes timeline-pulse {
        0% { box-shadow: 0 0 0 0 rgba(0, 212, 255, 0.7); }
        70% { box-shadow: 0 0 0 10px rgba(0, 212, 255, 0); }
        100% { box-shadow: 0 0 0 0 rgba(0, 212, 255, 0); }
    }
</style>
""", unsafe_allow_html=True)

# Import RFQ Modules
sys.path.insert(0, BASE_DIR)
try: import rfq_scraper; RFQ_AVAILABLE = True
except: RFQ_AVAILABLE = False
try: import rfq_submit; RFQ_SUBMIT_AVAILABLE = True
except: RFQ_SUBMIT_AVAILABLE = False
try: import rfq_fetch_prices; RFQ_FETCH_AVAILABLE = True
except: RFQ_FETCH_AVAILABLE = False

# --- DATA HELPERS ---
def load_config():
    cfg = {}
    if os.path.exists(CONFIG_PATH):
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            for line in f:
                if "=" in line:
                    k, v = line.split("=", 1)
                    cfg[k.strip()] = v.strip()
    return cfg

def save_config(new_config):
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        for k, v in new_config.items():
            f.write(f"{k}={v}\n")

def load_data():
    if os.path.exists(MASTER_FILE):
        df = pd.read_excel(MASTER_FILE)
        if 'Status' in df.columns:
            df['Status'] = df['Status'].fillna('UNKNOWN').apply(
                lambda x: '🟢 3. QUOTED' if 'REPLIED' in str(x).upper() else x
            )
        col_id = 'UID' if 'UID' in df.columns else ('MSG_ID' if 'MSG_ID' in df.columns else None)
        if col_id: df['UID'] = df[col_id].fillna(0).astype(str)
        else: df['UID'] = df.index.astype(str)
        return df.iloc[::-1].reset_index(drop=True)
    return None

def update_excel_row(uid, new_status):
    from openpyxl import load_workbook
    try:
        wb = load_workbook(MASTER_FILE); ws = wb.active; updated = False
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=11).value
            try:
                if float(val) == float(uid): ws.cell(row=r, column=2).value = new_status; updated = True
            except:
                if str(val).strip() == str(uid).strip(): ws.cell(row=r, column=2).value = new_status; updated = True
        if updated: wb.save(MASTER_FILE)
        wb.close(); return updated
    except: return False

# --- UI COMPONENTS ---
def login_screen():
    st.markdown('<div class="animated-title" style="text-align: center;">🔐 System Login</div>', unsafe_allow_html=True)
    with st.container():
        col_l, col_mid, col_r = st.columns([1, 2, 1])
        with col_mid:
            st.write("---")
            user = st.text_input("Username")
            pw = st.text_input("Password", type="password")
            if st.button("Login", use_container_width=True, type="primary"):
                success, info = auth_manager.check_login(user, pw)
                if success:
                    st.session_state['logged_in'] = True
                    st.session_state['user_info'] = info
                    st.success(f"Welcome, {info['name']}!")
                    time.sleep(0.5)
                    st.rerun()
                else:
                    st.error(info)

def main_dashboard(user_info):
    st.markdown(f'<div style="display: flex; align-items: center; gap: 20px;"><div class="animated-title">Order list Management</div> <div style="color: #00eb93; font-size: 0.9rem; font-weight: bold; border: 1px solid rgba(0,235,147,0.3); padding: 4px 12px; border-radius: 20px; display: flex; align-items: center; gap: 8px; background: rgba(0,235,147,0.05);"><div class="notif-dot"></div> ONLINE</div></div>', unsafe_allow_html=True)
    cfg = load_config()
    gsheet_enabled = cfg.get('ENABLE_GSHEET', 'True').upper() == 'TRUE'
    
    # Initialize navigation and notification state
    if 'active_page' not in st.session_state: st.session_state['active_page'] = 'Orders'
    if 'orders_notif' not in st.session_state: st.session_state['orders_notif'] = False
    if 'logs_notif' not in st.session_state: st.session_state['logs_notif'] = False
    
    # --- HELPER FUNCTIONS (Moved up to fix scope issues) ---
    def color_status(val):
        v = str(val).upper()
        if 'WAITING' in v: return 'background-color: rgba(255, 75, 75, 0.15); color: #ff4b4b;'
        if 'ACKNOWLEDGED' in v: return 'background-color: rgba(255, 165, 0, 0.15); color: #ffa500;'
        if 'QUOTED' in v: return 'background-color: rgba(0, 235, 147, 0.15); color: #00eb93;'
        return ''

    def color_rfq_status(val):
        v = str(val).upper()
        if 'DONE' in v or 'COMPLETED' in v: return 'background-color: rgba(0, 235, 147, 0.15); color: #00eb93;'
        if 'PENDING' in v or 'IN PROGRESS' in v: return 'background-color: rgba(255, 165, 0, 0.15); color: #ffa500;'
        if 'SUBMITTED' in v: return 'background-color: rgba(0, 212, 255, 0.15); color: #00d4ff;'
        return ''

    # --- INTERFACE UTILITIES ---
    def extract_customer_from_desc(desc):
        """Smart extractor to find customer names hidden in the RFQ description text."""
        d = str(desc)
        # Pattern 1: "Customer: Wai Mun / Wong"
        m1 = re.search(r"[Cc]ustomer:\s*([^▪|•]+)", d)
        if m1: return m1.group(1).strip()
        # Pattern 2: "WaiMun.Wong - • human head..."
        m2 = re.search(r"^([^-\s•][^-\s•]*?)\s*-\s*", d)
        if m2: return m2.group(1).strip()
        # Pattern 3: "Samsung... – Bin Othman, Amir Nazrin"
        m3 = re.search(r"–\s*([^–]+)$", d)
        if m3: return m3.group(1).strip()
        return "-"

    def render_status_pipeline(status):
        """Visual Status Pipeline for the Action Center."""
        s = str(status).upper()
        p_width = "0%"
        steps = ["", "", ""] # [Done/Active/Pending for each step]
        
        # Step 1: Created, Step 2: Processing (Waiting), Step 3: Quoted
        if 'WAITING' in s:
            p_width = "50%"
            steps = ["done", "active timeline-active-pulse", ""]
        elif 'QUOTED' in s or 'DONE' in s:
            p_width = "100%"
            steps = ["done", "done", "done"]
        else:
            p_width = "15%" # Created phase
            steps = ["active", "", ""]

        st.markdown(f"""
        <div class="timeline-container">
            <div class="timeline-line"></div>
            <div class="timeline-progress" style="width: {p_width};"></div>
            <div class="timeline-step {steps[0]}">
                <div class="timeline-dot">1</div>
                <div class="timeline-label">Order<br>Received</div>
            </div>
            <div class="timeline-step {steps[1]}">
                <div class="timeline-dot">2</div>
                <div class="timeline-label">Price<br>Fetching</div>
            </div>
            <div class="timeline-step {steps[2]}">
                <div class="timeline-dot">3</div>
                <div class="timeline-label">Quotation<br>Sent</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

    df = load_data()
    
    # 🕵️ Detect Status Updates via Data Fingerprinting
    if df is not None:
        # Create a hash of the current statuses to detect any backend changes
        status_fingerprint = str(df[['UID', 'Status']].values.tolist())
        if 'last_fingerprint' not in st.session_state:
            st.session_state['last_fingerprint'] = status_fingerprint
        elif status_fingerprint != st.session_state['last_fingerprint']:
            # Something changed in the Excel file!
            st.session_state['orders_notif'] = True
            st.session_state['logs_notif'] = True
            st.session_state['last_fingerprint'] = status_fingerprint
            st.session_state['show_dot'] = True # Sidebar dot too

    # Clear notifications for currently active page
    if st.session_state['active_page'] == 'Orders': st.session_state['orders_notif'] = False
    if st.session_state['active_page'] == 'Logs': st.session_state['logs_notif'] = False
    
    # Check if we should still show the sidebar dot (if any notification remains)
    if not st.session_state['orders_notif'] and not st.session_state['logs_notif']:
        st.session_state['show_dot'] = False

    # Sidebar: Info & Logout
    with st.sidebar:
        st.write(f"👤 **User:** {user_info['name']} ({user_info['role'].upper()})")
        if st.button("🚪 Logout", use_container_width=True):
            auth_manager.logout()
        
        if st.session_state.get('show_dot', False):
            st.markdown("---")
            st.write("🔔 **New Updates Detected**")
            if st.button("🔕 Reset Alert", use_container_width=True):
                st.session_state['show_dot'] = False
                st.session_state['notif_cleared'] = True
                st.rerun()
        st.markdown("---")

    # --- SMART NAVIGATION BAR (Replacing st.tabs) ---
    st.markdown("""
        <style>
            .nav-container { display: flex; gap: 10px; margin-bottom: 25px; border-bottom: 1px solid rgba(255,255,255,0.1); padding-bottom: 5px; }
            .nav-item { position: relative; display: flex; align-items: center; }
            .nav-pulse { position: absolute; top: -5px; right: -5px; }
        </style>
    """, unsafe_allow_html=True)

    c_nav1, c_nav2, c_nav3, c_nav4 = st.columns([1, 1, 1, 1])
    
    with c_nav1:
        o_btn_type = "primary" if st.session_state['active_page'] == 'Orders' else "secondary"
        if st.button(f"📊 Orders", use_container_width=True, type=o_btn_type):
            st.session_state['active_page'] = 'Orders'; st.rerun()
        if st.session_state['orders_notif']:
            st.markdown('<div style="position: relative;"><div class="notif-dot" style="position: absolute; top: -35px; right: 5px;"></div></div>', unsafe_allow_html=True)
            
    with c_nav2:
        l_btn_type = "primary" if st.session_state['active_page'] == 'Logs' else "secondary"
        if st.button(f"📝 RFQ Logs", use_container_width=True, type=l_btn_type):
            st.session_state['active_page'] = 'Logs'; st.rerun()
        if st.session_state['logs_notif']:
            st.markdown('<div style="position: relative;"><div class="notif-dot" style="position: absolute; top: -35px; right: 5px;"></div></div>', unsafe_allow_html=True)

    if user_info['role'] == 'admin':
        with c_nav3:
            if st.button("⚙️ Settings", use_container_width=True): st.session_state['active_page'] = 'Settings'; st.rerun()
        with c_nav4:
            if st.button("👥 Users", use_container_width=True): st.session_state['active_page'] = 'Users'; st.rerun()

    # --- PAGE RENDERING ---
    active_page = st.session_state['active_page']

    # --- PAGE: ORDERS ---
    if active_page == 'Orders':
        st.markdown(f"""
            <div style="display: flex; align-items: center; justify-content: space-between; gap: 10px; margin-bottom: 20px;">
                <h2 style="margin: 0;">📊 Orders</h2>
                <div style="color: #00eb93; font-size: 0.8rem; font-weight: bold; display: flex; align-items: center; gap: 6px;"><div class="notif-dot"></div> PORTAL CONNECTED</div>
            </div>
        """, unsafe_allow_html=True)
        
        if df is not None:
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Total", len(df))
            c2.metric("Waiting", len(df[df['Status'].str.contains('WAITING', na=False)]))
            c3.metric("Quoted", len(df[df['Status'].str.contains('QUOTED', na=False)]))
            
            st.markdown("---")
            f1, f2 = st.columns([1, 2])
            with f1:
                search = st.text_input("Search Customer/Item", "")
                show_del = st.checkbox("Show Deleted")
            with f2:
                all_st = df['Status'].unique().tolist()
                def_st = [s for s in all_st if 'DELETED' not in str(s).upper()]
                st_filter = st.multiselect("Status Filter", options=all_st, default=def_st)
            
            f_df = df[df['Status'].isin(st_filter)]
            if not show_del: f_df = f_df[~f_df['Status'].str.contains('DELETED', na=False)]
            if search: f_df = f_df[f_df['Customer Name'].str.contains(search, case=False, na=False) | f_df['Item Name'].str.contains(search, case=False, na=False)]
            
            # --- CUSTOMER NAME TO THE FRONT ---
            if 'Customer Name' in f_df.columns:
                cols = ['Customer Name'] + [c for c in f_df.columns if c != 'Customer Name']
                f_df = f_df[cols]

            st.dataframe(f_df.style.applymap(color_status, subset=['Status']), use_container_width=True, height=400)

            # --- Row Action (Redesigned for better focus) ---
            if not f_df.empty:
                st.markdown("---")
                st.subheader("⚡ Quick Action Center")
                
                sel_items = f_df.head(20).copy()
                sel_items['Selector'] = sel_items['UID'] + " | " + sel_items['Customer Name'].astype(str)
                
                col_sel, col_empty = st.columns([2, 1])
                with col_sel:
                    selected = st.selectbox("🎯 Step 1: Select Order to Action", options=sel_items['Selector'].tolist())
                
                if selected:
                    suid = str(selected.split(" | ")[0])
                    s_item = sel_items[sel_items['UID'] == suid].iloc[0]
                    
                    st.markdown(f"""
                        <div class="info-box">
                            <strong>📦 Target Item:</strong> {s_item['Customer Name']} — <span style="color: #00d4ff;">{s_item['Item Name']}</span>
                        </div>
                    """, unsafe_allow_html=True)
                    
                    # RENDER TIMELINE
                    render_status_pipeline(s_item['Status'])
                    
                    st.write("---")
                    st.subheader("⚙️ Step 2: Choose Action")
                    
                    sc1, sc2, sc3 = st.columns(3)
                    
                    with sc1:
                        with st.container(border=True):
                            st.markdown("#### 🟢 Status Update")
                            new_s = st.selectbox("Current Progress", ["🔴 1. WAITING", "🟡 2. ACKNOWLEDGED", "🟢 3. QUOTED", "⚪ 0. MANUAL REVIEW"])
                            if st.button("Apply New Status", use_container_width=True):
                                if update_excel_row(suid, new_s): st.success("Updated!"); time.sleep(0.5); st.rerun()

                    with sc2:
                        with st.container(border=True):
                            st.markdown("#### ☁️ Portal Submission")
                            if RFQ_SUBMIT_AVAILABLE:
                                draft_key = f"draft_path_{suid}"
                                
                                if st.button("🛠️ 1. Generate RFQ Draft", use_container_width=True):
                                    with st.spinner("Generating Excel..."):
                                        try:
                                            path = rfq_submit.generate_draft_excel(s_item.to_dict())
                                            st.session_state[draft_key] = path
                                            st.success(f"Generated! Saved in Extracted_Emails")
                                            time.sleep(0.5)
                                            st.rerun()
                                        except Exception as e:
                                            st.error(f"Generate error: {e}")
                                            
                                if st.session_state.get(draft_key):
                                    d_path = st.session_state[draft_key]
                                    if st.button("📂 2. Open Draft to Check", use_container_width=True):
                                        try:
                                            os.startfile(os.path.abspath(d_path))
                                        except Exception as e:
                                            st.error(f"Cannot auto-open: {e}")
                                            
                                    if st.button("📤 3. Confirm & Submit", type="primary", use_container_width=True):
                                        with st.spinner("Submitting checked Excel to portal..."):
                                            ok, msg = rfq_submit.submit_rfq(s_item.to_dict(), filepath=d_path)
                                            if ok:
                                                st.balloons()
                                                st.success(f"### ✅ {msg}")
                                                rfq_scraper.clear_cache(); # Instant Log Update
                                                st.session_state[draft_key] = None # Clear after submit
                                            else:
                                                st.error(f"### ❌ {msg}")

                    with sc3:
                        with st.container(border=True):
                            st.markdown("#### 💾 G-Sheet Sync")
                            if RFQ_FETCH_AVAILABLE:
                                rfq = st.text_input("Ref RFQ No.", key="sync_no", value=s_item.get('UID',''))
                                
                                # Pull Cost logic
                                if st.button("🔍 Pull Cost from Portal", use_container_width=True):
                                    import rfq_fetch_prices
                                    with st.spinner("Fetching..."):
                                        res = rfq_fetch_prices.fetch_rfq_cost(rfq)
                                        if res['success'] and res['file_found']:
                                            st.session_state['cost_input_val'] = res['total_cost']
                                            st.session_state['manual_cost'] = res['total_cost']
                                            if 'breakdown' in res:
                                                st.session_state['cost_breakdown'] = res['breakdown']
                                            
                                            # --- INJECT IMAGES INTO MASTER ORDERS ---
                                            if 'images' in res and len(res['images']) > 0:
                                                try:
                                                    ok_inj, msg_inj = rfq_fetch_prices.inject_images_to_master(rfq, res['images'], MASTER_FILE)
                                                    if ok_inj:
                                                        st.success(f"📸 {msg_inj} into Master_Orders!")
                                                    else:
                                                        st.warning(f"⚠️ Image Sync: {msg_inj}")
                                                except Exception as e:
                                                    st.error(f"⚠️ Need 'Pillow' library for images! Run: pip install Pillow")
                                            
                                            st.rerun() 
                                        else: st.warning(res['msg'])

                                cost_val = st.number_input("💵 Cost (MYR)", value=st.session_state.get('manual_cost', 0.0), key="cost_input_val")
                                
                                # --- Calculation Breakdown Display ---
                                if st.session_state.get('cost_breakdown'):
                                    with st.expander("🔢 View Calculation Breakdown", expanded=False):
                                        for item in st.session_state['cost_breakdown']:
                                            st.write(f"🔹 **{item['desc']}**")
                                            st.write(f"&nbsp;&nbsp;&nbsp;&nbsp;{item['qty']} units × {item['cost']:.2f} = **{item['subtotal']:.2f}**")
                                        st.divider()
                                        st.write(f"📝 **Total Landed Cost:** {st.session_state.get('manual_cost', 0.0):.2f}")
                                markup_val = st.number_input("📈 Markup (%)", value=20.0, step=1.0)
                                selling_val = cost_val * (1 + markup_val/100)
                                profit_val = selling_val - cost_val
                                st.info(f"**Selling:** {selling_val:.2f} | **Profit:** {profit_val:.2f}")

                                btn_label = "Sync Price & Log" if gsheet_enabled else "❌ Logic Disabled"
                                if st.button(btn_label, use_container_width=True, disabled=not gsheet_enabled, type="primary"):
                                    import gsheet_logger
                                    ok, msg = gsheet_logger.update_rfq_quotation(rfq, cost_val, selling_val, profit_val, markup_val, order_row=s_item.to_dict())
                                    if ok: 
                                        st.success(msg); 
                                        rfq_scraper.clear_cache(); # Instant Log Update
                                        st.balloons()
                                    else: st.error(msg)

    # --- PAGE: RFQ LOGS ---
    elif active_page == 'Logs':
        st.markdown(f"""
            <div style="display: flex; align-items: center; justify-content: space-between; gap: 10px; margin-bottom: 20px;">
                <h2 style="margin: 0;">📝 RFQ Portal Live Feed</h2>
                <div style="color: #00eb93; font-size: 0.8rem; font-weight: bold; display: flex; align-items: center; gap: 6px;"><div class="notif-dot"></div> SYSTEM ONLINE</div>
            </div>
        """, unsafe_allow_html=True)
        
        col_ref, col_spacer = st.columns([1, 2])
        with col_ref:
            if st.button("🔄 Fetch Latest RFQs", use_container_width=True, type="primary"):
                rfq_scraper.clear_cache()
                st.rerun()

        if RFQ_AVAILABLE:
            data, err = rfq_scraper.get_rfq_data()
            if data:
                rdf = pd.DataFrame(data)
                
                # --- FILTER BY CREATOR (Dynamic from config) ---
                portal_user = cfg.get('PORTAL_USER_NAME', 'Chee Kah Chun')
                if 'Created By' in rdf.columns:
                    rdf = rdf[rdf['Created By'].str.contains(portal_user, case=False, na=False)]
                
                # --- RFQ CUSTOMER EXTRACTION & PRIORITY ---
                if 'Item Desc.' in rdf.columns:
                    rdf.insert(0, 'Customer', rdf['Item Desc.'].apply(extract_customer_from_desc))
                
                # --- SMART FILTERING (Hide Testing/Ignored items) ---
                if not rdf.empty:
                    ignore_mask = rdf['Item Desc.'].str.contains('TEST|IGNORE|DELETE', case=False, na=False)
                    
                    # Target specific bad IDs directly across all columns just to be safe
                    targets = ['GX001378', 'GX001379']
                    id_mask = rdf.apply(lambda row: any(t in str(val) for val in row.values for t in targets), axis=1)
                    
                    rdf = rdf[~(ignore_mask | id_mask)]

                if not rdf.empty:
                    if 'Status' in rdf.columns:
                        st.dataframe(rdf.style.applymap(color_rfq_status, subset=['Status']), use_container_width=True)
                    else:
                        st.dataframe(rdf, use_container_width=True)
                else:
                    st.info("No active RFQs found (or all are filtered out).")
            elif err: st.error(err)

    # --- PAGE: SYSTEM SETTINGS (Admin Only) ---
    elif active_page == 'Settings' and user_info['role'] == 'admin':
        st.subheader("⚙️ System Configuration")
        config = load_config()
        with st.form("config_form"):
            new_cfg = {}
            new_cfg['IMAP_SERVER'] = st.text_input("IMAP Server", config.get('IMAP_SERVER',''))
            new_cfg['EMAIL_USER'] = st.text_input("Email Username", config.get('EMAIL_USER',''))
            new_cfg['EMAIL_PASSWORD'] = st.text_input("Email Password", config.get('EMAIL_PASSWORD',''), type="password")
            new_cfg['GEMINI_API_KEY'] = st.text_input("Gemini API Key", config.get('GEMINI_API_KEY',''), type="password")
            new_cfg['SPREADSHEET_ID'] = st.text_input("Google Sheet ID", config.get('SPREADSHEET_ID',''))
            new_cfg['ENABLE_GSHEET'] = str(st.checkbox("Enable Google Sheet Sync", value=(config.get('ENABLE_GSHEET','True').upper() == 'TRUE')))
            new_cfg['RFQ_URL'] = st.text_input("RFQ Portal URL", config.get('RFQ_URL',''))
            new_cfg['PORTAL_USER_NAME'] = st.text_input("Portal Creator Name (Filter)", config.get('PORTAL_USER_NAME','Chee Kah Chun'))
            
            # Keep hidden ones
            for k in ['IMAP_PORT', 'CHECK_INTERVAL_SECONDS', 'OUTPUT_DIR', 'RFQ_EMAIL', 'RFQ_PASSWORD']:
                new_cfg[k] = config.get(k, '')
                
            if st.form_submit_button("💾 Save All Config"):
                save_config(new_cfg)
                st.success("Config saved successfully!")

    # --- PAGE: USER MANAGEMENT (Admin Only) ---
    elif active_page == 'Users' and user_info['role'] == 'admin':
        st.subheader("👥 User Account Management")
        users = auth_manager.load_users()
        st.table(pd.DataFrame(users)[['username', 'name', 'role', 'is_active']])
        
        st.markdown("---")
        with st.expander("➕ Add New User"):
            with st.form("add_user"):
                nu = st.text_input("New Username")
                np = st.text_input("New Password", type="password")
                nn = st.text_input("Full Name")
                nr = st.selectbox("Role", ["user", "admin"])
                if st.form_submit_button("Create Account"):
                    ok, msg = auth_manager.add_user(nu, np, nn, nr)
                    if ok: st.success(msg); time.sleep(0.5); st.rerun()
                    else: st.error(msg)
        
        with st.expander("🔒 Toggle User Status"):
            target = st.selectbox("Select User", [u['username'] for u in users if u['username'] != user_info['username']])
            action = st.radio("Status", ["Enable", "Disable"])
            if st.button("Apply User Status"):
                if auth_manager.update_user_status(target, action == "Enable"):
                    st.success(f"User {target} updated."); time.sleep(0.5); st.rerun()

# --- MAIN ENTRY ---
auth_manager.init_session()
if not st.session_state['logged_in']:
    login_screen()
else:
    main_dashboard(st.session_state['user_info'])
